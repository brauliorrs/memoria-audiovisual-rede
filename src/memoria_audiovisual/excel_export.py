from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def autosize_worksheet(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def style_header(row):
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font


def append_sheet(workbook, title, rows, fieldnames):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(fieldnames)
    style_header(worksheet[1])
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
    autosize_worksheet(worksheet)
    return worksheet


def save_excel_report(
    path,
    payload,
    rows_summary,
    rows_video_links,
    rows_internal_pages,
    ccaaa_members,
    ccaaa_rows_summary,
    ccaaa_rows_video_links,
    ccaaa_rows_internal_pages,
):
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.append(["Metrica", "Valor"])
    style_header(dashboard[1])
    dashboard.append(["Fonte", payload["source"]])
    dashboard.append(["Total de instituicoes", payload["total_institutions"]])
    dashboard.append(["Sites OK", payload["ok_count"]])
    dashboard.append(["Com links de video", payload["with_video_count"]])
    dashboard.append(["Com midia embutida", payload["with_embeds_count"]])
    for status, count in payload["status_summary"].items():
        dashboard.append([f"Status: {status}", count])
    for status, count in payload["integrity_summary"].items():
        dashboard.append([f"Integridade: {status}", count])
    autosize_worksheet(dashboard)

    append_sheet(
        workbook,
        "Ranking",
        payload["top_institutions"],
        [
            "institution",
            "slug",
            "status",
            "integrity_status",
            "priority_review",
            "video_links_found_total",
            "embedded_video_signals_total",
            "partner_domain",
            "final_url",
        ],
    )
    append_sheet(
        workbook,
        "Summary",
        rows_summary,
        [
            "institution",
            "slug",
            "country",
            "continent",
            "partner_site",
            "partner_domain",
            "status",
            "http_code",
            "integrity_status",
            "final_url",
            "video_links_found_total",
            "embedded_video_signals_total",
            "candidate_internal_pages",
            "priority_review",
            "warning",
            "error",
        ],
    )
    append_sheet(
        workbook,
        "Video Links",
        rows_video_links,
        ["institution", "slug", "country", "continent", "partner_site", "platform", "video_link"],
    )
    append_sheet(
        workbook,
        "Internal Pages",
        rows_internal_pages,
        [
            "institution",
            "slug",
            "country",
            "continent",
            "partner_site",
            "internal_page",
            "status",
            "http_code",
            "video_links_found",
            "embedded_signals",
            "warning",
            "error",
        ],
    )
    append_sheet(
        workbook,
        "CCAAA Members",
        ccaaa_members,
        ["organization", "abbreviation", "role", "website", "domain", "description", "source"],
    )
    append_sheet(
        workbook,
        "CCAAA Summary",
        ccaaa_rows_summary,
        [
            "institution",
            "slug",
            "abbreviation",
            "role",
            "partner_site",
            "partner_domain",
            "status",
            "http_code",
            "integrity_status",
            "final_url",
            "video_links_found_total",
            "embedded_video_signals_total",
            "candidate_internal_pages",
            "priority_review",
            "warning",
            "member_source",
            "error",
        ],
    )
    append_sheet(
        workbook,
        "CCAAA Video Links",
        ccaaa_rows_video_links,
        [
            "institution",
            "slug",
            "abbreviation",
            "role",
            "partner_site",
            "platform",
            "video_link",
            "member_source",
        ],
    )
    append_sheet(
        workbook,
        "CCAAA Internal",
        ccaaa_rows_internal_pages,
        [
            "institution",
            "slug",
            "abbreviation",
            "role",
            "partner_site",
            "internal_page",
            "status",
            "http_code",
            "video_links_found",
            "embedded_signals",
            "warning",
            "member_source",
            "error",
        ],
    )

    workbook.save(path)
